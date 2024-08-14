from flask import Flask, render_template, request
import pandas as pd
import category_encoders as ce
from sklearn.naive_bayes import GaussianNB
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
import openpyxl
app = Flask(__name__)

# Đọc dữ liệu từ file Excel
df = pd.read_excel("./data//dataNCKH.xlsx")
# df.head()
# Đặt tên cột
col_names = ['symptom_1', 'symptom_2', 'symptom_3', 'symptom_4', 'diagnose']
df.columns = col_names

# Lấy biến phân loại
categorical = ['symptom_1', 'symptom_2', 'symptom_3', 'symptom_4']

# Tạo encoder
encoder = ce.OneHotEncoder(cols=categorical)

# Thực hiện one-hot encoding trên toàn bộ dữ liệu
df_encoded = encoder.fit_transform(df)

# Tạo và huấn luyện mô hình Naive Bayes
X = df_encoded.drop(['diagnose'], axis=1)
y = df_encoded['diagnose']

# Chia dữ liệu thành tập huấn luyện và tập kiểm tra
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)


# from sklearn.naive_bayes import MultinomialNB

# mnb = MultinomialNB()
# mnb.fit(X_train, y_train)

# from sklearn.naive_bayes import BernoulliNB

# bnb = BernoulliNB()
# bnb.fit(X_train, y_train)

# from sklearn.naive_bayes import ComplementNB

# cnb = ComplementNB()
# cnb.fit(X_train, y_train)

# from sklearn.naive_bayes import CategoricalNB

# cat_nb = CategoricalNB()
# cat_nb.fit(X_train, y_train)


gnb = GaussianNB()
gnb.fit(X_train, y_train)

    
# Định nghĩa route cho trang chủ
@app.route('/')
def index():
    return render_template('index.html')

# Định nghĩa route cho việc nhận dữ liệu từ form và thực hiện dự đoán
@app.route('/predict', methods=['POST'])
def predict():
    symptoms = [request.form['symptom_1'], request.form['symptom_2'],
                request.form['symptom_3'], request.form['symptom_4']]
    
    # Lấy cột 'diagnose' từ DataFrame gốc
    user_data = df[['symptom_1', 'symptom_2', 'symptom_3', 'symptom_4', 'diagnose']].copy()
    
    # Thêm dữ liệu người dùng vào DataFrame
    user_data.loc[len(user_data)] = [symptoms[0], symptoms[1], symptoms[2], symptoms[3], None]
    
    # Áp dụng encoder đã học từ toàn bộ dữ liệu
    user_data_encoded = encoder.transform(user_data)

    # Dự đoán bằng mô hình đã huấn luyện
    prediction = gnb.predict(user_data_encoded.iloc[-1, :-1].values.reshape(1, -1))

    # Tính toán độ chính xác, điểm training và điểm test
    accuracy = accuracy_score(y_test, gnb.predict(X_test)) * 100
    training_score = gnb.score(X_train, y_train) * 100
    test_score = gnb.score(X_test, y_test) * 100
    
    accuracy = round(accuracy, 2)
    training_score = round(training_score, 2)
    test_score = round(test_score, 2)

    return render_template('result.html', prediction=prediction[0], accuracy=accuracy, training_score=training_score, test_score=test_score)


@app.route('/save-feedback', methods=['POST'])
def save_feedback():
    email = request.form.get('email')
    feedback = request.form.get('feedback')
    # Đọc dữ liệu từ file Excel (feedback.xlsx)
    workbook = openpyxl.load_workbook('./data/feedback.xlsx')
    sheet = workbook.active
    # Tìm hàng cuối cùng trong sheet
    last_row = sheet.max_row
    next_row = last_row + 1
    # Thêm dữ liệu mới vào sheet
    sheet.cell(row=next_row, column=1, value=email)
    sheet.cell(row=next_row, column=2, value=feedback)
    # Ghi lại dữ liệu vào file Excel
    workbook.save('./data/feedback.xlsx')
    return render_template('index.html')


@app.route('/add-data', methods=['POST'])
def save_data():
    email = request.form.get('email')
    symptom_1 =request.form.get('symptom_1')
    symptom_2 =request.form.get('symptom_2')
    symptom_3 =request.form.get('symptom_3')
    symptom_4 =request.form.get('symptom_4')
    diagnose = request.form.get('diagnose')
    # Đọc dữ liệu từ file Excel (feedback.xlsx)
    workbook = openpyxl.load_workbook('./data/user_data.xlsx')
    sheet = workbook.active
    # Tìm hàng cuối cùng trong sheet
    last_row = sheet.max_row
    next_row = last_row + 1
    # Thêm dữ liệu mới vào sheet
    sheet.cell(row=next_row, column=1, value=email)
    sheet.cell(row=next_row, column=2, value=symptom_1)
    sheet.cell(row=next_row, column=3, value=symptom_2)
    sheet.cell(row=next_row, column=4, value=symptom_3)
    sheet.cell(row=next_row, column=5, value=symptom_4)
    sheet.cell(row=next_row, column=6, value=diagnose)
    # Ghi lại dữ liệu vào file Excel
    workbook.save('./data/user_data.xlsx')
    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
